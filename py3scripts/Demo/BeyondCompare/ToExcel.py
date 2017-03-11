# -*- coding:utf-8 -*-
import os
import sys
from openpyxl import load_workbook

def outputTree(tree, level, result):
    files = []
    level += 1
    for key in sorted(tree.keys()):
        if type(tree[key]) == int:
            files.append(key)
        else:
            result.append(['']*level)
            result[-1].append(key)
            result[-1].insert(0, -1)
            outputTree(tree[key], level, result)
    for item in files:
        result.append(['']*level)
        result[-1].append(item)
        result[-1].insert(0, tree[item])
def drawTree(result, startpos, stoppos, modidx):
    mark1  = r'┣━'
    mark2  = r'┃  '
    mark3  = r'┗━'
    if startpos >= len(result):
        return
    curlevel = len(result[startpos])
    if curlevel <= (modidx + 1):
        return
    pos=[startpos]
    for i in range(startpos + 1, stoppos):
        if len(result[i]) == curlevel:
            pos.append(i)
    for i in range(startpos, pos[-1]):
        if i in pos:
            result[i][modidx] = mark1
        else:
            result[i][modidx] = mark2
    result[pos[-1]][modidx] = mark3
    pos.append(stoppos)
    for i in range(len(pos)-1):
        drawTree(result, pos[i]+1, pos[i+1], modidx + 1)

def toExcel(rootdir, outxlsm, outxlsx, tmpl, excludes=[]):
    htmllist = []
    excludelist = [os.path.abspath(x) for x in excludes]
    rootpath = os.path.abspath(rootdir)
    # walk directory
    strippos = len(rootpath) + 1
    for root,dirs,files in os.walk(rootpath):
        for item in files:
            if os.path.splitext(item)[1].lower() == '.html':
                htmlpath = os.path.join(root, item)
                if htmlpath not in excludelist:
                    htmlurl = 'file:///{}'.format(htmlpath.replace(os.sep, '/'))
                    srcname = os.path.splitext(item)[0]
                    htmllist.append([htmlurl, htmlpath[strippos:], srcname])
    # generate sheetname for each diff worksheet
    maxNameLen = 31
    nameList = {}
    for idx, item in enumerate(htmllist):
        item[2] = item[2][:maxNameLen]
        nameList.setdefault(item[2], []).append(idx)
    for item in nameList.keys():
        nameCount = len(nameList[item])
        if nameCount > 1:
            nameCount = len(str(nameCount))
            pat = "{0}_{1:0" + str(nameCount) + "}"
            for subidx, subitem in enumerate(nameList[item]):
                htmllist[subitem][2] = pat.format(htmllist[subitem][2][:-(nameCount+1)], subidx)
    del nameList
    # generate file tree
    tree = {}
    for idx, item in enumerate(htmllist):
        parts = item[1].split(os.sep)
        worktree = tree
        for subidx, subitem in enumerate(parts):
            if subidx >= len(parts) - 1:
                worktree[subitem] = idx
            else:
                worktree = worktree.setdefault(subitem, {})
    result = []
    outputTree(tree, 0, result)
    drawTree(result, 0, len(result), 1)
    # output xlsm
    wb = load_workbook(tmpl, keep_vba=True)
    ws = wb.active
    ws['E1'] = os.path.abspath(outxlsx)
    for idx,item in enumerate(result):
        for subidx in range(1, len(item)-1):
            ws.cell(column=subidx+3, row=idx+2).value=item[subidx]
        if item[0] >= 0:
            ws.cell(column=1, row=idx+2).value=htmllist[item[0]][0]
            ws.cell(column=2, row=idx+2).value=len(item) + 2
            ws.cell(column=3, row=idx+2).value=htmllist[item[0]][2]
            ws.cell(column=len(item)+2, row=idx+2).value=os.path.splitext(item[-1])[0]
        else:
            ws.cell(column=len(item)+2, row=idx+2).value=item[-1]
    wb.save(outxlsm)

if __name__ == '__main__':
    rootdir = '.'
    tmpl = r'd:\temp\FolderCompare\tmpl.xlsm'
    outxlsm = r'd:\temp\FolderCompare\test.xlsm'
    outxlsx = r'd:\temp\FolderCompare\test.xlsx'
    toExcel(rootdir, outxlsm, outxlsx, tmpl, [os.path.join(rootdir, 'Report.html')])
